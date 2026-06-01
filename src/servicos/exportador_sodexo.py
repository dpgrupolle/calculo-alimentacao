"""
Preenche a planilha da operadora (Sodexo), respeitando as regras do Erick:

- Mexe SÓ na aba "Dados dos Beneficiários-Cartão". As outras ficam intactas.
- Única coluna alterada: "Valor crédito" (= cálculo do benefício).
- Preenche também "Data de crédito", "Data de entrega" e "Mês de referência"
  (informadas pelo usuário antes de gerar).
- Zerados → fonte vermelha na linha.
- "Atestado" escrito na coluna ao lado SÓ para quem zerou por causa de atestado.
- Exceção (diretoria/sócios): se o valor que já está na planilha não é um valor
  de faixa, mantém como está (não recalcula nem desconta).

Devolve os bytes do .xlsx preenchido.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font

from src.servicos.matching import so_digitos, tira_um
from src.utils.formatadores import normalizar_busca

ABA = "Dados dos Beneficiários-Cartão"
VERMELHO = "FFFF0000"
PRETO = "FF000000"


def _achar_aba(wb):
    # tolerante a variações no nome
    for nome in wb.sheetnames:
        n = normalizar_busca(nome)
        if "beneficiari" in n and "cart" in n:
            return wb[nome]
    raise ValueError("Não achei a aba 'Dados dos Beneficiários-Cartão' na planilha Sodexo.")


def _mapear_colunas(ws):
    """Acha a linha de cabeçalho e devolve {chave: índice de coluna (1-based)}."""
    alvo = {
        "matricula": {"matricula", "matrícula"},
        "cpf": {"cpf"},
        "valor": {"valor credito", "valor crédito"},
        "data_credito": {"data de credito", "data de crédito"},
        "data_entrega": {"data de entrega"},
        "mes_ref": {"mes de referencia", "mês de referência", "mes de referência"},
    }
    for r in range(1, 15):
        achados = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            nv = normalizar_busca(str(v))
            for chave, apelidos in alvo.items():
                if nv in apelidos:
                    achados[chave] = c
        if "matricula" in achados and "valor" in achados:
            achados["cabecalho_linha"] = r
            # coluna de anotação "Atestado" = logo após Mês de referência
            base = achados.get("mes_ref", achados["valor"])
            achados["anotacao"] = base + 1
            return achados
    raise ValueError("Não achei o cabeçalho (Matrícula / Valor crédito) na aba Sodexo.")


def _to_dt(d):
    if isinstance(d, (datetime,)):
        return d
    if isinstance(d, date):
        return datetime(d.year, d.month, d.day)
    return d


def preencher_sodexo(template_bytes: bytes, resultados: list, faixa_valor: dict,
                     data_credito, data_entrega, mes_referencia) -> dict:
    """
    template_bytes: bytes do .xlsx modelo (guardado/enviado).
    resultados: lista do motor de cálculo (calcular_processo).
    faixa_valor: {faixa: valor} — para detectar exceções (valor fora da tabela).
    datas: date/datetime para crédito, entrega e mês de referência.
    Retorna {"bytes":..., "preenchidos":n, "zerados":n, "excecoes":n,
             "sem_match":[...], "avisos":[...]}.
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = _achar_aba(wb)
    col = _mapear_colunas(ws)
    lin0 = col["cabecalho_linha"] + 1

    # índice de resultados por código (Sankhya) e por CPF
    por_cod = {}
    por_cpf = {}
    for r in resultados:
        if r.get("codigo"):
            por_cod[str(r["codigo"]).strip()] = r
        if r.get("cpf"):
            por_cpf[so_digitos(r["cpf"])] = r

    valores_faixa = set(round(float(v), 2) for v in faixa_valor.values() if v not in (None, 0))

    dt_cred = _to_dt(data_credito)
    dt_ent = _to_dt(data_entrega)
    dt_mes = _to_dt(mes_referencia)

    preenchidos = zerados = excecoes = 0
    sem_match = []
    avisos = []

    r = lin0
    while r <= ws.max_row:
        mat = ws.cell(r, col["matricula"]).value
        if mat in (None, ""):
            r += 1
            # para no primeiro bloco vazio após dados
            # (algumas planilhas têm linhas extras em branco no fim)
            # verifica se ainda há matrícula abaixo
            resto = any(ws.cell(rr, col["matricula"]).value not in (None, "")
                        for rr in range(r, min(r + 5, ws.max_row + 1)))
            if not resto:
                break
            continue

        mats = str(mat).strip()
        # remove ".0" caso venha numérico
        if mats.endswith(".0"):
            mats = mats[:-2]
        cpf = so_digitos(ws.cell(r, col["cpf"]).value) if col.get("cpf") else ""

        # localiza resultado: código direto → CPF → tira-1 (este por último,
        # pois "tira o 1" de matrículas curtas pode colidir, ex.: "11" → "1")
        res = (por_cod.get(mats)
               or (por_cpf.get(cpf) if cpf else None)
               or por_cod.get(tira_um(mats)))

        valor_atual = ws.cell(r, col["valor"]).value
        eh_excecao = False
        try:
            va = round(float(valor_atual), 2)
            if va != 0 and va not in valores_faixa:
                eh_excecao = True
        except (TypeError, ValueError):
            va = None

        cel_valor = ws.cell(r, col["valor"])
        cel_anot = ws.cell(r, col["anotacao"])

        if eh_excecao:
            # mantém o valor que já está (diretoria/sócios)
            excecoes += 1
            valor_final = va
            zerado = False
            por_ferias = False
            por_atestado = False
        elif res is not None:
            valor_final = res["valor_final"]
            cel_valor.value = valor_final
            zerado = (valor_final == 0)
            por_ferias = bool(res.get("zerado_por_ferias"))
            por_atestado = bool(res.get("zerado_por_atestado"))
            preenchidos += 1
        else:
            # não achou no cálculo — mantém o que está e avisa
            sem_match.append(mats)
            valor_final = va if va is not None else valor_atual
            zerado = (va == 0)
            por_ferias = False
            por_atestado = False

        # datas (em todas as linhas de dados)
        if col.get("data_credito"):
            ws.cell(r, col["data_credito"]).value = dt_cred
        if col.get("data_entrega"):
            ws.cell(r, col["data_entrega"]).value = dt_ent
        if col.get("mes_ref"):
            ws.cell(r, col["mes_ref"]).value = dt_mes

        # cor: zerado → vermelho; senão → preto (limpa marca antiga)
        cor = VERMELHO if zerado else PRETO
        for c in range(col["matricula"], col["anotacao"] + 1):
            cell = ws.cell(r, c)
            f = cell.font
            cell.font = Font(name=f.name, size=f.size, bold=f.bold,
                             italic=f.italic, color=cor)
        if zerado:
            zerados += 1

        # anotação ao lado: "Férias" tem prioridade; senão "Atestado"; senão limpa
        if por_ferias:
            cel_anot.value = "Férias"
        elif por_atestado:
            cel_anot.value = "Atestado"
        else:
            cel_anot.value = None

        r += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    if sem_match:
        avisos.append(
            f"{len(sem_match)} beneficiário(s) da planilha Sodexo não foram "
            f"encontrados no cálculo (mantive o valor que estava). Ex: {sem_match[:6]}"
        )
    return {
        "bytes": buffer.getvalue(),
        "preenchidos": preenchidos,
        "zerados": zerados,
        "excecoes": excecoes,
        "sem_match": sem_match,
        "avisos": avisos,
    }
