"""Exportador interno: resultado do processo em .xlsx (conferência), colorido por faixa."""
from __future__ import annotations
import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILL_INTEGRAL = PatternFill("solid", fgColor="D4EDDA")
FILL_PARCIAL = PatternFill("solid", fgColor="FFF3CD")
FILL_ZERO = PatternFill("solid", fgColor="F8D7DA")
FILL_CAB = PatternFill("solid", fgColor="041747")
FONT_CAB = Font(color="FFFFFF", bold=True)
BORDA = Border(*([Side(style="thin", color="D9D9D9")] * 4))


def _fill(pct):
    if pct >= 100: return FILL_INTEGRAL
    if pct <= 0: return FILL_ZERO
    return FILL_PARCIAL


def gerar_xlsx_resultado(processo: dict, resultados: List[dict]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Resultado"
    cols = [("Código", "codigo", 12), ("Nome", "nome", 30), ("Cargo", "cargo", 22),
            ("Faixa", "faixa_alim", 8), ("Prêmio", "valor_base", 14), ("% Desc.", "pct_desconto", 10),
            ("% Receb.", "pct_recebido", 10), ("Valor final", "valor_final", 14), ("Motivo", "motivo", 40)]
    mes = processo.get("mes_referencia", "")
    ws.append([f"Benefício — {mes}"]); ws["A1"].font = Font(bold=True, size=13, color="041747")
    ws.append([])
    hr = ws.max_row + 1
    for i, (t, _, w) in enumerate(cols, 1):
        c = ws.cell(hr, i, t); c.fill = FILL_CAB; c.font = FONT_CAB
        c.alignment = Alignment(horizontal="center"); c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in resultados:
        ws.append([r.get(k) for _, k, _ in cols])
        ln = ws.max_row; f = _fill(float(r.get("pct_recebido", 100)))
        for i, (_, k, _) in enumerate(cols, 1):
            cell = ws.cell(ln, i); cell.fill = f; cell.border = BORDA
            if k in ("valor_base", "valor_final"): cell.number_format = 'R$ #,##0.00'
            if k in ("pct_desconto", "pct_recebido"): cell.number_format = '0"%"'
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
