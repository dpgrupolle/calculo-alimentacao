"""Tela Modelo Sodexo (Opção C): guarda a planilha-base e permite atualizar."""
from __future__ import annotations
import openpyxl, io
import streamlit as st
from src.banco import repo_sodexo
from src.servicos import auditoria
from src.utils.formatadores import formatar_data
from src.utils.permissoes import pode_editar
from src.utils.marca import EMPRESAS


def _conta_beneficiarios(conteudo: bytes) -> int:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
        for nome in wb.sheetnames:
            n = nome.lower()
            if "beneficiári" in n and "cart" in n or "beneficiarios" in n and "cart" in n:
                ws = wb[nome]
                cnt = 0
                for row in ws.iter_rows(min_row=8, max_col=5):
                    if row and len(row) >= 4 and row[3].value not in (None, ""):
                        cnt += 1
                return cnt
    except Exception:
        pass
    return 0


def renderizar_modelo_sodexo(usuario):
    st.markdown("## 📤 Modelo Sodexo")
    st.caption("A planilha-padrão da operadora fica guardada aqui. Só muda quando entra/sai funcionário.")
    pode = pode_editar(usuario)

    msg = st.session_state.pop("sdx_msg", None)
    if msg:
        {"sucesso": st.success, "erro": st.error, "aviso": st.warning}.get(msg["tipo"], st.info)(msg["texto"])

    empresa = st.radio("Empresa", EMPRESAS, horizontal=True, key="sdx_emp")
    ativo = repo_sodexo.modelo_ativo(empresa)
    if ativo:
        st.success(f"✅ Modelo de **{empresa}**: {ativo['nome_arquivo']}  \n"
                   f"{ativo['total_beneficiarios']} beneficiário(s) · enviado por "
                   f"{ativo['enviado_por_nome'] or '—'} em {formatar_data(ativo['enviado_em'])}")
    else:
        st.warning(f"Nenhum modelo guardado para **{empresa}** ainda. Suba a planilha abaixo.")

    if not pode:
        return
    st.markdown("#### Subir / atualizar modelo")
    st.caption("Use quando entrar/sair funcionário. A versão nova substitui a anterior como base.")
    arq = st.file_uploader(f"Planilha da Sodexo de {empresa} (.xlsx)", type=["xlsx"], key=f"upl_sdx_{empresa}")
    if arq is not None and st.button(f"💾 Guardar como modelo de {empresa}", type="primary"):
        try:
            conteudo = arq.getvalue()
            total = _conta_beneficiarios(conteudo)
            repo_sodexo.salvar_modelo(arq.name, conteudo, total, usuario, empresa)
            auditoria.registrar(usuario, "GERAR_PLANILHA_SAIDA", detalhes=f"Modelo Sodexo: {arq.name}")
            st.session_state["sdx_msg"] = {"tipo": "sucesso", "texto": f"✅ Modelo guardado ({total} beneficiários)."}
        except Exception as e:
            st.session_state["sdx_msg"] = {"tipo": "erro", "texto": f"❌ {e}"}
        st.rerun()
